import streamlit as st
import io
import json
import re
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# --- Page Config ---
st.set_page_config(
    page_title="Azure NSG Rules Converter",
    page_icon="🎈",
    layout="wide"
)

# --- Header Section with Azure Branding ---
st.markdown("""
    <div style="display:flex; align-items:center; gap:15px; background-color:#0078D4; padding:15px; border-radius:10px; color:white;">
        <img src="https://upload.wikimedia.org/wikipedia/commons/f/fa/Microsoft_Azure.svg" width="60">
        <h1 style="margin:0;">Azure NSG Rules - JSON to Excel Converter</h1>
    </div>
    <p style="margin-top:10px; font-size:16px; color:#333;">
        Upload an Azure Network Security Group (NSG) JSON export and convert it into a clean, formatted Excel report 📊
    </p>
""", unsafe_allow_html=True)

# --- Region formatter ---
def format_location(loc: str) -> str:
    """Convert Azure location code into a human-friendly region name."""
    if not loc:
        return ""
    loc_lower = loc.lower()
    region_map = {
        # --- Australia / APAC ---
        "australiaeast": "Australia East",
        "australiasoutheast": "Australia Southeast",
        "australiacentral": "Australia Central",
        "australiacentral2": "Australia Central 2",
        "southeastasia": "Southeast Asia",
        "eastasia": "East Asia",
        "japaneast": "Japan East",
        "japanwest": "Japan West",
        "koreacentral": "Korea Central",
        "koreasouth": "Korea South",
        "southindia": "South India",
        "centralindia": "Central India",
        "westindia": "West India",

        # --- China (21Vianet) ---
        "chinanorth": "China North",
        "chinanorth2": "China North 2",
        "chinaeast": "China East",
        "chinaeast2": "China East 2",

        # --- Europe ---
        "northeurope": "North Europe",
        "westeurope": "West Europe",
        "francecentral": "France Central",
        "francesouth": "France South",
        "germanynorth": "Germany North",
        "germanywestcentral": "Germany West Central",
        "norwayeast": "Norway East",
        "norwaywest": "Norway West",
        "swedencentral": "Sweden Central",
        "swedensouth": "Sweden South",
        "switzerlandnorth": "Switzerland North",
        "switzerlandwest": "Switzerland West",
        "polandcentral": "Poland Central",
        "italynorth": "Italy North",
        "spaincentral": "Spain Central",
        "ukwest": "UK West",
        "uksouth": "UK South",

        # --- Americas ---
        "eastus": "East US",
        "eastus2": "East US 2",
        "westus": "West US",
        "westus2": "West US 2",
        "westus3": "West US 3",
        "centralus": "Central US",
        "northcentralus": "North Central US",
        "southcentralus": "South Central US",
        "westcentralus": "West Central US",
        "canadacentral": "Canada Central",
        "canadaeast": "Canada East",
        "brazilsouth": "Brazil South",
        "brazilsoutheast": "Brazil Southeast",
        "mexicocentral": "Mexico Central",
        "chilecentral": "Chile Central",

        # --- Middle East / Africa ---
        "uaecentral": "UAE Central",
        "uaenorth": "UAE North",
        "qatarcentral": "Qatar Central",
        "southafricanorth": "South Africa North",
        "southafricawest": "South Africa West",
        "israelcentral": "Israel Central",

        # --- US Government / DoD ---
        "usgovvirginia": "US Gov Virginia",
        "usgovarizona": "US Gov Arizona",
        "usgoviowa": "US Gov Iowa",
        "usgovtexas": "US Gov Texas",
        "usdodeast": "US DoD East",
        "usdodcentral": "US DoD Central",

        # --- Special / Edge ---
        "global": "Global",
        "centraluseuap": "Central US EUAP",
        "eastus2euap": "East US 2 EUAP"
    }

    if loc_lower in region_map:
        return region_map[loc_lower]
    # fallback for unknown regions
    loc_cleaned = re.sub(r'([a-z])([A-Z0-9])', r'\1 \2', loc_lower.title())
    loc_cleaned = loc_cleaned.replace("Azure ", "").replace("-", " ").title()
    return loc_cleaned

# --- File Upload Section ---
uploaded_file = st.file_uploader("📂 Upload your Azure NSG JSON file", type="json")

if uploaded_file is not None:
    # --- Parse JSON ---
    data = json.load(uploaded_file)
    nsg_name = data.get("name", "Unknown NSG")
    id_path = data.get("id", "")
    subscription_id = id_path.split("/")[2] if id_path else ""
    resource_group = (
        id_path.split("/resourceGroups/")[1].split("/")[0]
        if "/resourceGroups/" in id_path else ""
    )
    metadata = {
        "Resource Group": resource_group,
        "Location": format_location(data.get("location", "")),
        "Subscription ID": subscription_id,
    }

    # --- NSG Rules ---
    rules = data["properties"].get("securityRules", []) + data["properties"].get("defaultSecurityRules", [])

    def replace_any(v):
        if isinstance(v, str):
            return "Any" if v.strip() == "*" else v
        if isinstance(v, list):
            return ", ".join(replace_any(x) for x in v if x)
        return v

    records = []
    for r in rules:
        p = r.get("properties", {})
        ports = p.get("destinationPortRanges", []) or [p.get("destinationPortRange", "Any")]
        sources = p.get("sourceAddressPrefixes", []) or [p.get("sourceAddressPrefix", "Any")]
        dests = p.get("destinationAddressPrefixes", []) or [p.get("destinationAddressPrefix", "Any")]
        records.append({
            "Priority": int(p.get("priority", 0)),
            "Direction": p.get("direction", ""),
            "Rule Name": r.get("name", ""),
            "Port": replace_any(ports),
            "Protocol": replace_any(p.get("protocol", "")),
            "Source": replace_any(sources),
            "Destination": replace_any(dests),
            "Access": p.get("access", ""),
            "Description": p.get("description", "")
        })

    df_rules = pd.DataFrame(records)
    if not df_rules.empty:
        df_rules["Direction"] = pd.Categorical(df_rules["Direction"], categories=["Inbound", "Outbound"], ordered=True)
        df_rules = df_rules.sort_values(["Direction", "Priority"])

    # --- Excel File Creation ---
    wb = Workbook()
    ws = wb.active
    ws.title = "NSG_RULES"

    # Styles
    bold = Font(bold=True)
    title_font = Font(bold=True, size=14)
    hdr_fill = PatternFill("solid", fgColor="D9E1F2")
    title_fill = PatternFill("solid", fgColor="BDD7EE")
    align_center = Alignment(horizontal="center")
    border = Border(left=Side(style="thin"), right=Side(style="thin"),
                    top=Side(style="thin"), bottom=Side(style="thin"))

    # Header Info
    ws.append([nsg_name])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
    ws["A1"].font = title_font
    ws["A1"].fill = title_fill
    ws["A1"].alignment = align_center
    ws.append([""])
    for k, v in metadata.items():
        ws.append([k, v])
        ws[f"A{ws.max_row}"].font = bold
    ws.append([""])

    # Rules Table
    ws.append(["NSG RULES"])
    ws.append(df_rules.columns.tolist())
    for i in range(1, len(df_rules.columns) + 1):
        cell = ws[f"{get_column_letter(i)}{ws.max_row}"]
        cell.font = bold
        cell.fill = hdr_fill
    for r in df_rules.itertuples(index=False):
        ws.append(list(r))

    # Borders & Widths
    for row in ws.iter_rows():
        for c in row:
            if c.value:
                c.border = border
    for col in ws.columns:
        ws.column_dimensions[get_column_letter(col[0].column)].width = max(len(str(c.value)) for c in col if c.value) + 3

    # Save to memory
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    # --- Display on Streamlit ---
    st.markdown("### 📘 NSG Metadata")
    st.table(pd.DataFrame(list(metadata.items()), columns=["Field", "Value"]))

    st.markdown("### 🔒 NSG Rules Overview")
    st.dataframe(df_rules, use_container_width=True)

    # --- Download Button ---
    st.download_button(
        label="💾 Download Excel Report",
        data=excel_file,
        file_name=f"{nsg_name}_NSG_Rules.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        help="Click to download the formatted Excel report."
    )

else:
    st.info("👆 Please upload a JSON file to begin.")

# --- Footer ---
st.markdown("""
    <hr style="margin-top:40px;">
    <div style="text-align:center; color:gray; font-size:14px;">
        Built with ❤️ using Streamlit & Microsoft Azure<br>
        © 2025 Hashim Hilal — Cloud Architect 
    </div>
""", unsafe_allow_html=True)
